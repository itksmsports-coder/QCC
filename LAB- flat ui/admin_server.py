"""
QC Control Center — Backend Server
===================================
Handles live department data, chart logging, audit logs, Excel exports,
AND the full report archive system:

  reports/
    active/current.json       ← live chart_log (replaces chart_log.json)
    daily/YYYY-MM-DD.json
    weekly/YYYY-week-WW.json
    monthly/YYYY-MM.json
    annual/YYYY.json

At 06:00 AM the server automatically:
  1. Archives the current day → daily JSON
  2. Appends that day to the current weekly / monthly / annual JSONs
  3. Rolls over weekly/monthly/annual files when a new period starts
  4. Resets the active chart log for the new day

All /api/report/* endpoints read the archived files and return structured
JSON that the existing frontend already knows how to render.
"""

import os
import json
import threading
from datetime import datetime, date, timedelta
from flask import Flask, jsonify, request, send_from_directory

# ──────────────────────────────────────────────────────────────────────────────
# APP SETUP
# ──────────────────────────────────────────────────────────────────────────────

app = Flask(__name__)

BASE_DIR     = os.path.dirname(os.path.abspath(__file__))
REPORTS_DIR  = os.path.join(BASE_DIR, "reports")
ACTIVE_DIR   = os.path.join(REPORTS_DIR, "active")
DAILY_DIR    = os.path.join(REPORTS_DIR, "daily")
WEEKLY_DIR   = os.path.join(REPORTS_DIR, "weekly")
MONTHLY_DIR  = os.path.join(REPORTS_DIR, "monthly")
ANNUAL_DIR   = os.path.join(REPORTS_DIR, "annual")

# Active chart log lives inside reports/active/ so archive logic can swap it
CHART_LOG_PATH = os.path.join(ACTIVE_DIR, "current.json")
AUDIT_LOG_PATH = os.path.join(BASE_DIR, "audit_log.json")
AUDIT_LOG_MAX  = 1000

# Thread lock — prevents concurrent writes corrupting JSON files
_file_lock = threading.Lock()

# ──────────────────────────────────────────────────────────────────────────────
# DEPARTMENT CONFIGURATION
# ──────────────────────────────────────────────────────────────────────────────

DEPARTMENTS = {
    "CUTTING": os.path.join(BASE_DIR, "CUTTING", "data.json"),
    "EMBRO":   os.path.join(BASE_DIR, "EMBRO",   "data.json"),
}

DEFAULT_DEFECTS_CUTTING = {
    "WRONG_COLOR": 0, "WRONG_SIZE": 0, "WRONG_CUT": 0,
    "OFF_SHADE": 0, "LOW_QUALITY": 0, "WRONG_MATERIAL": 0, "OTHERS": 0,
}
DEFAULT_DEFECTS_EMBRO = {
    "LOW_QUALITY": 0, "WRONG_COLOR_THREAD": 0, "WRONG_PLACEMENT": 0,
    "WRONG_TEXT_FONT": 0, "UNBALANCED": 0, "WRONG_SPELLING": 0, "OTHERS": 0,
}
DEFAULT_DATA_CUTTING = {"employees": [
    {"name": "Employee 1", "defects": dict(DEFAULT_DEFECTS_CUTTING)},
    {"name": "Employee 2", "defects": dict(DEFAULT_DEFECTS_CUTTING)},
    {"name": "Employee 3", "defects": dict(DEFAULT_DEFECTS_CUTTING)},
]}
DEFAULT_DATA_EMBRO = {"employees": [
    {"name": "Employee 1", "defects": dict(DEFAULT_DEFECTS_EMBRO)},
    {"name": "Employee 2", "defects": dict(DEFAULT_DEFECTS_EMBRO)},
    {"name": "Employee 3", "defects": dict(DEFAULT_DEFECTS_EMBRO)},
]}
DEPT_DEFAULT_DATA = {"CUTTING": DEFAULT_DATA_CUTTING, "EMBRO": DEFAULT_DATA_EMBRO}

# ──────────────────────────────────────────────────────────────────────────────
# BOOTSTRAP — create all required folders and data files
# ──────────────────────────────────────────────────────────────────────────────

def bootstrap():
    """Create folder tree and initial data files if missing."""
    for d in [ACTIVE_DIR, DAILY_DIR, WEEKLY_DIR, MONTHLY_DIR, ANNUAL_DIR]:
        os.makedirs(d, exist_ok=True)

    for dept, path in DEPARTMENTS.items():
        os.makedirs(os.path.dirname(path), exist_ok=True)
        if not os.path.exists(path):
            with open(path, "w", encoding="utf-8") as f:
                json.dump(DEPT_DEFAULT_DATA.get(dept, DEFAULT_DATA_EMBRO), f, indent=2)
            print(f"[BOOTSTRAP] Created {path}")
        else:
            print(f"[OK] {dept} -> {path}")

    print(f"[OK] Report dirs ready under {REPORTS_DIR}")

# ──────────────────────────────────────────────────────────────────────────────
# GENERIC FILE HELPERS
# ──────────────────────────────────────────────────────────────────────────────

def _read_json(path, default=None):
    """Thread-safe JSON read.  Returns `default` if file missing or corrupt."""
    try:
        with _file_lock:
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
    except FileNotFoundError:
        return default if default is not None else {}
    except Exception as e:
        print(f"[READ ERROR] {path}: {e}")
        return default if default is not None else {}


def _write_json(path, data):
    """Thread-safe atomic JSON write (write temp → rename)."""
    tmp = path + ".tmp"
    try:
        with _file_lock:
            os.makedirs(os.path.dirname(path), exist_ok=True)
            with open(tmp, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=2)
            os.replace(tmp, path)
        return True
    except Exception as e:
        print(f"[WRITE ERROR] {path}: {e}")
        try:
            os.remove(tmp)
        except Exception:
            pass
        return False

# ──────────────────────────────────────────────────────────────────────────────
# EMPLOYEE / DEFECT HELPERS
# ──────────────────────────────────────────────────────────────────────────────

def _normalize_employees(employees):
    """Uppercase all defect keys; clamp values to >= 0."""
    if not isinstance(employees, list):
        return employees
    out = []
    for emp in employees:
        if not isinstance(emp, dict):
            out.append(emp)
            continue
        emp = dict(emp)
        defects = emp.get("defects", {})
        if isinstance(defects, dict):
            emp["defects"] = {
                k.upper(): max(0, int(v or 0))
                for k, v in defects.items()
            }
        out.append(emp)
    return out


def _normalize_defects(defects):
    if not isinstance(defects, dict):
        return {}
    return {str(k).upper(): max(0, int(v or 0)) for k, v in defects.items()}


def load_json(path):
    data = _read_json(path, {"employees": []})
    if isinstance(data, dict) and "employees" in data:
        data["employees"] = _normalize_employees(data["employees"])
    return data


def save_json(path, data):
    return _write_json(path, data)

# ──────────────────────────────────────────────────────────────────────────────
# CHART LOG (active/current.json)
# ──────────────────────────────────────────────────────────────────────────────

def _reset_boundary_for(dt: datetime) -> datetime:
    """Return the 06:00 AM boundary that governs the current day's data."""
    boundary = datetime(dt.year, dt.month, dt.day, 6, 0, 0)
    if dt.hour < 6:
        boundary -= timedelta(days=1)
    return boundary


def load_chart_log():
    """
    Load the active chart log.
    If the stored last_reset is older than the current 06:00 AM boundary,
    trigger a day-end archive and reset automatically.
    """
    data = _read_json(CHART_LOG_PATH, {})

    now           = datetime.now()
    reset_boundary = _reset_boundary_for(now)

    last_reset_str = data.get("last_reset", "")
    needs_reset    = True

    if last_reset_str:
        try:
            last_reset_dt = datetime.fromisoformat(last_reset_str)
            if last_reset_dt >= reset_boundary:
                needs_reset = False
        except Exception:
            pass

    if needs_reset:
        # Archive yesterday's data before wiping
        if data.get("entries"):
            _do_archive(data, reset_boundary)

        data = {
            "last_reset":  reset_boundary.isoformat(),
            "reset_date":  reset_boundary.date().isoformat(),
            "entries":     [],
        }
        _write_json(CHART_LOG_PATH, data)
        print(f"[CHART LOG] Reset at 6AM boundary ({reset_boundary})")

    return data


def save_chart_log(data):
    return _write_json(CHART_LOG_PATH, data)


def append_chart_entry(department, total_defects):
    data = load_chart_log()
    now  = datetime.now()
    data.setdefault("entries", []).append({
        "timestamp":    now.isoformat(),
        "hour_fraction": round(now.hour + now.minute / 60 + now.second / 3600, 4),
        "department":   department,
        "total":        total_defects,
    })
    save_chart_log(data)

# ──────────────────────────────────────────────────────────────────────────────
# ARCHIVE ENGINE
# ──────────────────────────────────────────────────────────────────────────────

def _iso_week(d: date):
    """Return (iso_year, iso_week) for a date object."""
    iso = d.isocalendar()
    return iso[0], iso[1]   # (year, week)


def _build_daily_snapshot(log_data: dict, snapshot_date: date) -> dict:
    """
    Aggregate a day's chart-log entries into a structured daily report.
    Reads the live department data files for employee/defect detail.
    """
    entries = log_data.get("entries", [])

    # Per-department aggregation
    dept_totals:  dict[str, int]        = {}   # dept -> max total seen (last entry)
    dept_entries: dict[str, list]       = {}   # dept -> all chart entries

    for e in entries:
        dept = e.get("department", "UNKNOWN")
        dept_entries.setdefault(dept, []).append(e)
        dept_totals[dept] = max(dept_totals.get(dept, 0), e.get("total", 0))

    # Build per-dept detail from the live data files
    departments: dict[str, dict] = {}
    for dept, path in DEPARTMENTS.items():
        live_data  = load_json(path)
        employees  = live_data.get("employees", [])
        defects_by_cat: dict[str, int] = {}
        emp_rows = []
        for emp in employees:
            d = emp.get("defects", {})
            row_total = sum(d.values())
            emp_rows.append({
                "name":    emp.get("name", "?"),
                "defects": dict(d),
                "total":   row_total,
            })
            for cat, val in d.items():
                defects_by_cat[cat] = defects_by_cat.get(cat, 0) + val

        total_dept  = sum(v for v in defects_by_cat.values())
        departments[dept] = {
            "total_defects":   total_dept,
            "employee_count":  len(employees),
            "employees":       emp_rows,
            "defects_by_cat":  defects_by_cat,
        }

    total_plant = sum(d["total_defects"] for d in departments.values())
    worst_dept  = max(departments, key=lambda k: departments[k]["total_defects"], default="—")

    return {
        "ok":          True,
        "date":        snapshot_date.isoformat(),
        "period":      "daily",
        "archived_at": datetime.now().isoformat(),
        "summary": {
            "total_plant_defects": total_plant,
            "worst_department":    worst_dept,
            "department_count":    len(departments),
        },
        "departments":   departments,
        "chart_entries": entries,
    }


def _do_archive(log_data: dict, boundary: datetime):
    """
    Called when the 6AM boundary rolls over.
    Writes daily JSON and appends to weekly/monthly/annual files.
    """
    # The archived day is the day BEFORE the new boundary date
    archived_date = (boundary - timedelta(seconds=1)).date()
    print(f"[ARCHIVE] Archiving {archived_date} …")

    # ── 1. Daily file ──────────────────────────────────────────────────────
    daily_path    = os.path.join(DAILY_DIR, f"{archived_date.isoformat()}.json")
    daily_report  = _build_daily_snapshot(log_data, archived_date)

    if not _write_json(daily_path, daily_report):
        print(f"[ARCHIVE] ERROR writing daily file {daily_path}")
        return

    print(f"[ARCHIVE] Daily → {daily_path}")

    # ── 2. Weekly file ─────────────────────────────────────────────────────
    iso_year, iso_week = _iso_week(archived_date)
    weekly_label = f"{iso_year}-week-{str(iso_week).zfill(2)}"
    weekly_path  = os.path.join(WEEKLY_DIR, f"{weekly_label}.json")
    _append_to_weekly(weekly_path, weekly_label, archived_date, daily_report)

    # ── 3. Monthly file ────────────────────────────────────────────────────
    monthly_label = f"{archived_date.year}-{str(archived_date.month).zfill(2)}"
    monthly_path  = os.path.join(MONTHLY_DIR, f"{monthly_label}.json")
    _append_to_monthly(monthly_path, monthly_label, archived_date, daily_report)

    # ── 4. Annual file ─────────────────────────────────────────────────────
    annual_path = os.path.join(ANNUAL_DIR, f"{archived_date.year}.json")
    _append_to_annual(annual_path, str(archived_date.year), archived_date, daily_report)

    print(f"[ARCHIVE] Done for {archived_date}.")


def _append_to_weekly(path: str, label: str, day: date, daily: dict):
    """Append a daily snapshot to the weekly report, rebuilding the summary."""
    data = _read_json(path, {
        "ok":    True,
        "label": label,
        "period": "weekly",
        "days":  {},
        "summary": {},
    })
    data["ok"]     = True
    data["label"]  = label
    data["period"] = "weekly"

    # Store the day's condensed data
    data.setdefault("days", {})[day.isoformat()] = {
        "date":               day.isoformat(),
        "total_plant_defects": daily["summary"]["total_plant_defects"],
        "worst_department":   daily["summary"]["worst_department"],
        "departments": {
            dept: v["total_defects"]
            for dept, v in daily.get("departments", {}).items()
        },
    }

    data["summary"] = _calc_weekly_summary(data["days"])
    _write_json(path, data)
    print(f"[ARCHIVE] Weekly → {path}")


def _calc_weekly_summary(days: dict) -> dict:
    if not days:
        return {}

    totals    = {d: v["total_plant_defects"] for d, v in days.items()}
    dept_sums: dict[str, int] = {}
    for v in days.values():
        for dept, cnt in (v.get("departments") or {}).items():
            dept_sums[dept] = dept_sums.get(dept, 0) + cnt

    total      = sum(totals.values())
    peak_day   = max(totals, key=totals.get)
    avg_daily  = round(total / len(totals), 1) if totals else 0

    return {
        "total_errors":         total,
        "errors_per_department": dept_sums,
        "average_daily_errors": avg_daily,
        "peak_error_day":       peak_day,
        "peak_error_count":     totals[peak_day],
    }


def _append_to_monthly(path: str, label: str, day: date, daily: dict):
    """Append a daily snapshot to the monthly report, rebuilding the summary."""
    data = _read_json(path, {
        "ok":     True,
        "label":  label,
        "period": "monthly",
        "days":   {},
        "summary": {},
    })
    data["ok"]     = True
    data["label"]  = label
    data["period"] = "monthly"

    data.setdefault("days", {})[day.isoformat()] = {
        "date":                day.isoformat(),
        "total_plant_defects": daily["summary"]["total_plant_defects"],
        "worst_department":    daily["summary"]["worst_department"],
        "departments": {
            dept: v["total_defects"]
            for dept, v in daily.get("departments", {}).items()
        },
    }

    data["summary"] = _calc_period_summary(data["days"])
    _write_json(path, data)
    print(f"[ARCHIVE] Monthly → {path}")


def _calc_period_summary(days: dict) -> dict:
    """Generic summary for weekly/monthly data (same structure)."""
    return _calc_weekly_summary(days)   # same logic


def _append_to_annual(path: str, label: str, day: date, daily: dict):
    """Append a daily snapshot to the annual report, rebuilding the summary."""
    data = _read_json(path, {
        "ok":     True,
        "label":  label,
        "period": "annual",
        "months": {},
        "summary": {},
    })
    data["ok"]     = True
    data["label"]  = label
    data["period"] = "annual"

    month_key   = f"{day.year}-{str(day.month).zfill(2)}"
    month_label = day.strftime("%B %Y")

    months = data.setdefault("months", {})
    if month_key not in months:
        months[month_key] = {
            "month_key":    month_key,
            "month_label":  month_label,
            "total_errors": 0,
            "departments":  {},
            "day_count":    0,
        }

    m = months[month_key]
    m["total_errors"] += daily["summary"]["total_plant_defects"]
    m["day_count"]    += 1

    for dept, v in daily.get("departments", {}).items():
        m["departments"][dept] = m["departments"].get(dept, 0) + v["total_defects"]

    data["summary"] = _calc_annual_summary(months)
    _write_json(path, data)
    print(f"[ARCHIVE] Annual → {path}")


def _calc_annual_summary(months: dict) -> dict:
    if not months:
        return {}

    totals     = {k: v["total_errors"] for k, v in months.items()}
    dept_sums: dict[str, int] = {}
    for v in months.values():
        for dept, cnt in (v.get("departments") or {}).items():
            dept_sums[dept] = dept_sums.get(dept, 0) + cnt

    total        = sum(totals.values())
    month_count  = len(totals)
    worst_month  = max(totals, key=totals.get)
    best_month   = min(totals, key=totals.get)
    avg_monthly  = round(total / month_count, 1) if month_count else 0

    dept_ranking = sorted(
        [{"department": d, "total": t} for d, t in dept_sums.items()],
        key=lambda x: x["total"],
        reverse=True,
    )

    return {
        "total_yearly_errors":    total,
        "errors_per_department":  dept_sums,
        "department_ranking":     dept_ranking,
        "average_monthly_errors": avg_monthly,
        "worst_month":            worst_month,
        "best_month":             best_month,
    }

# ──────────────────────────────────────────────────────────────────────────────
# AUDIT LOG
# ──────────────────────────────────────────────────────────────────────────────

def load_audit_log():
    data = _read_json(AUDIT_LOG_PATH, {"entries": []})
    if isinstance(data, dict) and isinstance(data.get("entries"), list):
        return data
    return {"entries": []}


def save_audit_log(data):
    entries = data.get("entries", [])
    if len(entries) > AUDIT_LOG_MAX:
        data["entries"] = entries[-AUDIT_LOG_MAX:]
    return _write_json(AUDIT_LOG_PATH, data)


def compute_employee_changes(before_emps, after_emps):
    before_emps = before_emps if isinstance(before_emps, list) else []
    after_emps  = after_emps  if isinstance(after_emps, list)  else []
    changes = []

    for i in range(len(after_emps), len(before_emps)):
        changes.append({"type": "employee_removed", "employee": before_emps[i].get("name","?")})
    for i in range(len(before_emps), len(after_emps)):
        changes.append({"type": "employee_added",   "employee": after_emps[i].get("name","?")})

    for i in range(min(len(before_emps), len(after_emps))):
        b, a = before_emps[i], after_emps[i]
        if b.get("name") != a.get("name"):
            changes.append({"type": "employee_renamed", "from": b.get("name","?"), "to": a.get("name","?")})
        bd = _normalize_defects(b.get("defects", {}))
        ad = _normalize_defects(a.get("defects", {}))
        diffs = {k: {"from": bd.get(k,0), "to": ad.get(k,0)}
                 for k in set(bd) | set(ad) if bd.get(k,0) != ad.get(k,0)}
        if diffs:
            changes.append({"type": "defects_updated", "employee": a.get("name","?"), "changes": diffs})

    return changes


def summarize_changes(action, changes):
    if action == "reset":
        prev = (changes[0].get("previous_total") if changes else 0)
        return f"Reset from {prev} → 0"
    if not changes:
        return "No changes"
    parts = []
    for item in changes:
        kind = item.get("type","")
        if kind == "employee_added":
            parts.append("added " + item.get("employee","?"))
        elif kind == "employee_removed":
            parts.append("removed " + item.get("employee","?"))
        elif kind == "employee_renamed":
            parts.append(item.get("from","?") + " → " + item.get("to","?"))
        elif kind == "defects_updated":
            parts.append(item.get("employee","?") + ": " + str(len(item.get("changes",{})or{})) + " field(s)")
    return "; ".join(parts) if parts else "Changes saved"


def append_audit_entry(department, action, user, changes):
    data  = load_audit_log()
    entry = {
        "timestamp":  datetime.now().isoformat(),
        "user":       str(user or "unknown").strip() or "unknown",
        "department": department,
        "action":     action,
        "changes":    changes,
        "summary":    summarize_changes(action, changes),
    }
    data.setdefault("entries", []).append(entry)
    save_audit_log(data)
    return entry

# ──────────────────────────────────────────────────────────────────────────────
# ROUTES — static files
# ──────────────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    resp = send_from_directory(BASE_DIR, "admin_dashboard.html")
    resp.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    return resp

@app.route("/logo.png")
def logo():
    resp = send_from_directory(BASE_DIR, "logo.png")
    resp.headers["Cache-Control"] = "public, max-age=86400"
    return resp

@app.route("/favicon.ico")
def favicon():
    return "", 204

@app.route("/users.json")
def users():
    return send_from_directory(BASE_DIR, "users.json")

# ──────────────────────────────────────────────────────────────────────────────
# ROUTES — live data
# ──────────────────────────────────────────────────────────────────────────────

@app.route("/api/departments")
def departments():
    return jsonify({dept: load_json(path) for dept, path in DEPARTMENTS.items()})


@app.route("/api/chart-log")
def get_chart_log():
    return jsonify(load_chart_log())


@app.route("/api/audit-log")
def get_audit_log():
    data    = load_audit_log()
    entries = list(reversed(data.get("entries", [])))
    return jsonify({"entries": entries})


@app.route("/api/update", methods=["POST"])
def update_department():
    try:
        body       = request.get_json(force=True)
        department = body.get("department")
        employees  = body.get("employees")
        updated_by = (body.get("updated_by") or body.get("username") or "").strip()

        if department not in DEPARTMENTS:
            return jsonify({"ok": False, "error": "Department not found"}), 404

        path          = DEPARTMENTS[department]
        existing_data = load_json(path)
        before_emps   = existing_data.get("employees", [])
        employees     = _normalize_employees(employees)
        changes       = compute_employee_changes(before_emps, employees)

        existing_data["employees"]       = employees
        existing_data["last_updated"]    = datetime.now().isoformat()
        existing_data["last_updated_by"] = updated_by

        if not save_json(path, existing_data):
            return jsonify({"ok": False, "error": "Save failed"}), 500

        total_defects = sum(v for emp in employees for v in emp.get("defects", {}).values())
        append_chart_entry(department, total_defects)
        append_audit_entry(department, "update", updated_by, changes)

        return jsonify({"ok": True})
    except Exception as e:
        print("Update Error:", e)
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/export", methods=["POST"])
def export_department():
    try:
        body       = request.get_json(force=True)
        department = body.get("department")
        if department not in DEPARTMENTS:
            return jsonify({"ok": False, "error": "Department not found"}), 404

        from io import BytesIO
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            return jsonify({"ok": False, "error": "openpyxl not installed. Run: pip install openpyxl"}), 500

        data       = load_json(DEPARTMENTS[department])
        employees  = data.get("employees", [])
        if not employees:
            return jsonify({"ok": False, "error": "No data to export"}), 400

        categories = list(employees[0].get("defects", {}).keys())
        wb  = Workbook()
        ws  = wb.active
        ws.title = department

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
        ws.append(["Employee"] + categories + ["Total"])
        for cell in ws[1]:
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="center")

        for emp in employees:
            row_total = sum(emp["defects"].get(c, 0) for c in categories)
            ws.append([emp["name"]] + [emp["defects"].get(c, 0) for c in categories] + [row_total])

        grand_total = sum(sum(emp["defects"].get(c,0) for c in categories) for emp in employees)
        ws.append(["GRAND TOTAL"] + [""] * len(categories) + [grand_total])
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)

        for col_cells in ws.columns:
            values = [str(c.value) for c in col_cells if c.value is not None]
            width  = max([len(v) for v in values] + [10])
            ws.column_dimensions[col_cells[0].column_letter].width = width + 4

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        from flask import Response
        filename = f"{department}_defects_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        resp = Response(buf.getvalue(),
                        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
        return resp
    except Exception as e:
        print("Export Error:", e)
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/reset", methods=["POST"])
def reset_department():
    try:
        body       = request.get_json(force=True)
        department = body.get("department")
        updated_by = (body.get("updated_by") or body.get("username") or "").strip()

        if department not in DEPARTMENTS:
            return jsonify({"ok": False, "error": "Department not found"}), 404

        path          = DEPARTMENTS[department]
        existing_data = load_json(path)
        before_total  = sum(v for emp in existing_data.get("employees", [])
                            for v in emp.get("defects", {}).values())

        for emp in existing_data.get("employees", []):
            emp["defects"] = {k: 0 for k in emp.get("defects", {})}

        existing_data["last_updated"]    = datetime.now().isoformat()
        existing_data["last_updated_by"] = updated_by

        if not save_json(path, existing_data):
            return jsonify({"ok": False, "error": "Save failed"}), 500

        append_chart_entry(department, 0)
        append_audit_entry(department, "reset", updated_by,
                           [{"type": "reset", "previous_total": before_total, "new_total": 0}])
        return jsonify({"ok": True})
    except Exception as e:
        print("Reset Error:", e)
        return jsonify({"ok": False, "error": str(e)}), 500

# ──────────────────────────────────────────────────────────────────────────────
# ROUTES — report API
# ──────────────────────────────────────────────────────────────────────────────

@app.route("/api/report/available-dates")
def available_dates():
    """
    Return a list of all dates that have a daily report file.
    The frontend calendar highlights these dates.
    """
    try:
        dates = []
        if os.path.isdir(DAILY_DIR):
            for fname in sorted(os.listdir(DAILY_DIR)):
                if fname.endswith(".json"):
                    dates.append(fname[:-5])   # strip .json → YYYY-MM-DD
        return jsonify({"ok": True, "dates": dates})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/report/daily/<date_str>")
def report_daily(date_str: str):
    """
    Return a daily report for YYYY-MM-DD.
    Special value 'today' returns a live-computed snapshot of the active log.
    """
    # Validate / parse
    if date_str == "today":
        live = load_chart_log()
        snapshot = _build_daily_snapshot(live, date.today())
        return jsonify(snapshot)

    try:
        parsed = date.fromisoformat(date_str)
    except ValueError:
        return jsonify({"ok": False, "error": "Invalid date format. Use YYYY-MM-DD."}), 400

    # Check if it's today (return live data)
    if parsed == date.today():
        live = load_chart_log()
        snapshot = _build_daily_snapshot(live, parsed)
        return jsonify(snapshot)

    path = os.path.join(DAILY_DIR, f"{date_str}.json")
    if not os.path.exists(path):
        return jsonify({"ok": False, "error": f"No report found for {date_str}."}), 404

    data = _read_json(path)
    data["ok"] = True
    return jsonify(data)


@app.route("/api/report/weekly/<week_label>")
def report_weekly(week_label: str):
    """
    Return weekly report.  Label format: YYYY-week-WW
    Special value 'current' returns the current ISO week.
    """
    if week_label == "current":
        today = date.today()
        iso_year, iso_week = _iso_week(today)
        week_label = f"{iso_year}-week-{str(iso_week).zfill(2)}"

    path = os.path.join(WEEKLY_DIR, f"{week_label}.json")
    if not os.path.exists(path):
        return jsonify({"ok": False, "error": f"No weekly report found for {week_label}."}), 404

    data = _read_json(path)
    data["ok"] = True
    return jsonify(data)


@app.route("/api/report/monthly/<month_label>")
def report_monthly(month_label: str):
    """
    Return monthly report.  Label format: YYYY-MM
    Special value 'current' returns the current month.
    """
    if month_label == "current":
        today = date.today()
        month_label = f"{today.year}-{str(today.month).zfill(2)}"

    path = os.path.join(MONTHLY_DIR, f"{month_label}.json")
    if not os.path.exists(path):
        return jsonify({"ok": False, "error": f"No monthly report found for {month_label}."}), 404

    data = _read_json(path)
    data["ok"] = True
    return jsonify(data)


@app.route("/api/report/annual/<year_label>")
def report_annual(year_label: str):
    """
    Return annual report.  Label format: YYYY
    Special value 'current' returns the current year.
    """
    if year_label == "current":
        year_label = str(date.today().year)

    path = os.path.join(ANNUAL_DIR, f"{year_label}.json")
    if not os.path.exists(path):
        return jsonify({"ok": False, "error": f"No annual report found for {year_label}."}), 404

    data = _read_json(path)
    data["ok"] = True
    return jsonify(data)


# ── Manual archive trigger (for testing / admin use) ─────────────────────────

@app.route("/api/report/archive-now", methods=["POST"])
def archive_now():
    """
    Manually trigger the day-end archive for the current active log.
    Useful for testing without waiting until 6AM.
    Requires a JSON body: { "target_date": "YYYY-MM-DD" }  (optional)
    """
    try:
        body = request.get_json(force=True) or {}
        target_date_str = body.get("target_date")
        if target_date_str:
            try:
                target = date.fromisoformat(target_date_str)
            except ValueError:
                return jsonify({"ok": False, "error": "Invalid target_date"}), 400
            # Build boundary as midnight of target+1 (to capture full day)
            boundary = datetime(target.year, target.month, target.day, 6, 0, 0) + timedelta(days=1)
        else:
            target   = date.today()
            boundary = datetime.now()

        live_data = load_chart_log()
        if not live_data.get("entries"):
            return jsonify({"ok": False, "error": "No chart entries to archive."}), 400

        _do_archive(live_data, boundary)
        return jsonify({"ok": True, "archived_date": target.isoformat()})
    except Exception as e:
        print("Archive-Now Error:", e)
        return jsonify({"ok": False, "error": str(e)}), 500

# ──────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ──────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print()
    print("===================================")
    print("  QC CONTROL CENTER — SERVER")
    print("===================================")
    bootstrap()
    print()
    print("Open your browser at:  http://localhost:5050")
    print()
    app.run(host="0.0.0.0", port=5050, debug=True, threaded=True)
