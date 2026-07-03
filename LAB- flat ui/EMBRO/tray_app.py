#!/usr/bin/env python3
"""
================================================================================
 EMBROIDERING DEFECT MONITORING BOARD  –  System Tray Launcher
 
 Requirements:
     pip install flask openpyxl pystray pillow

 HOW TO RUN (no console window):
     pythonw tray_app.py          ← Windows: use pythonw so no CMD appears
     -- OR --
     Build with PyInstaller:
         pyinstaller --noconsole --onefile tray_app.py

 The tray icon right-click menu offers:
     • Open Dashboard   – opens http://localhost:4646 in the browser
     • Restart Service  – stops and restarts the Flask server thread
     • Settings         – opens the admin panel URL
     • Stop Service     – shuts down Flask but keeps tray alive
     • Quit             – stops everything and exits
================================================================================
"""

import os
import sys
import threading
import webbrowser
from datetime import datetime
from io import BytesIO

# ---------------------------------------------------------------------------
# Graceful import checks
# ---------------------------------------------------------------------------
try:
    import pystray
    from pystray import MenuItem as item, Menu
except ImportError:
    sys.exit(
        "pystray is not installed.\n"
        "Run:  pip install pystray pillow\n"
        "Then re-launch."
    )

try:
    from PIL import Image, ImageDraw, ImageFont
except ImportError:
    sys.exit(
        "Pillow is not installed.\n"
        "Run:  pip install pillow\n"
        "Then re-launch."
    )

try:
    from flask import Flask, Response, jsonify, request, send_from_directory
except ImportError:
    sys.exit(
        "Flask is not installed.\n"
        "Run:  pip install flask\n"
        "Then re-launch."
    )

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

import json
import winreg
from threading import RLock
from werkzeug.serving import make_server

# ==============================================================================
# CONFIGURATION  (same as server.py)
# ==============================================================================

BASE_DIR        = os.path.dirname(os.path.abspath(__file__))
DATA_FILE       = os.path.join(BASE_DIR, "data.json")

# Folder layout:
#   Main folder/
#     CUTTING FOLDER/
#     EMBRO FOLDER/     <- this script lives here (BASE_DIR)
#     HTML REPORTS/     <- reports are saved here (shared by both apps)
REPORTS_DIR     = os.path.join(os.path.dirname(BASE_DIR), "HTML")
ADMIN_PASSWORD  = "KSM_RND*2026"
PORT            = 4646
DASHBOARD_URL   = f"http://localhost:{PORT}"

DEFECT_CATEGORIES = [
    {"key": "low_quality",          "label": "Low Quality",          "bg": "#FFD600", "fg": "#101820"},
    {"key": "wrong_color_thread",   "label": "Wrong Color Thread",   "bg": "#CE93D8", "fg": "#1a001f"},
    {"key": "wrong_placement",      "label": "Wrong Placement",      "bg": "#00E676", "fg": "#06210f"},
    {"key": "wrong_text_font",      "label": "Wrong Text Font",      "bg": "#42A5F5", "fg": "#0a1628"},
    {"key": "unbalanced",           "label": "Unbalanced",           "bg": "#FFB74D", "fg": "#2A1300"},
    {"key": "wrong_spelling",       "label": "Wrong Spelling",       "bg": "#FF1744", "fg": "#2a0008"},
    {"key": "others",               "label": "Others",               "bg": "#26C6DA", "fg": "#062024"},
]

THRESHOLD_ORANGE = 6
THRESHOLD_RED    = 16

# Maps any case-variant of a defect-category key (e.g. "WRONG_SPELLING",
# "Wrong_Spelling") back to our canonical lowercase key. The QC Control
# Center admin tool may write defect keys in a different case than we do;
# without this, those values silently read back as 0.
_KEY_LOOKUP = {c["key"].lower(): c["key"] for c in DEFECT_CATEGORIES}

# Top-level keys we understand. Anything else found in data.json (e.g.
# fields written by the QC Control Center admin tool) is treated as
# foreign data we must round-trip untouched rather than silently drop.
_KNOWN_TOP_KEYS = {"employees", "last_updated", "schedule"}

# Uptime tracking – reset each time the service starts
_service_start_time: datetime | None = None

# Windows Registry key for startup
_STARTUP_REG_KEY  = r"Software\Microsoft\Windows\CurrentVersion\Run"
_STARTUP_APP_NAME = "EmbroideryDefectMonitorBoard"

# ==============================================================================
# DATA HELPERS  (identical to server.py)
# ==============================================================================

_data_lock = RLock()


def _blank_defects():
    return {c["key"]: 0 for c in DEFECT_CATEGORIES}


def _sample_employees(count=12):
    return [{"name": f"Operator {i}", "defects": _blank_defects()}
            for i in range(1, count + 1)]


def _normalize_employee(raw):
    raw_defects = raw.get("defects") if isinstance(raw.get("defects"), dict) else {}
    # Case-insensitive lookup: the QC Control Center admin tool may write
    # this dict with differently-cased keys (e.g. "WRONG_SPELLING") than
    # our own canonical lowercase keys, so match on lowercase regardless
    # of what casing the value came in as.
    ci_defects = {str(k).lower(): v for k, v in raw_defects.items()}
    normalized = _blank_defects()
    for cat in DEFECT_CATEGORIES:
        try:
            normalized[cat["key"]] = max(0, int(ci_defects.get(cat["key"].lower(), 0)))
        except (TypeError, ValueError):
            normalized[cat["key"]] = 0
    return {
        "name": str(raw.get("name") or "Operator").strip() or "Operator",
        "defects": normalized,
    }


def _default_schedule():
    return {"hour": -1, "minute": 0}   # hour = -1 means OFF


def _normalize_schedule(raw):
    sched = raw if isinstance(raw, dict) else {}
    try:
        hour = int(sched.get("hour", -1))
    except (TypeError, ValueError):
        hour = -1
    try:
        minute = int(sched.get("minute", 0))
    except (TypeError, ValueError):
        minute = 0
    if hour < -1 or hour > 23:
        hour = -1
    minute = max(0, min(59, minute))
    return {"hour": hour, "minute": minute}


def migrate_state(loaded):
    # Anything in the file we don't explicitly model gets carried forward
    # untouched, so we never clobber fields the QC Control Center admin
    # tool owns just because we don't know about them.
    extra = {k: v for k, v in loaded.items()
             if k not in _KNOWN_TOP_KEYS and k != "departments"}

    schedule = _normalize_schedule(loaded.get("schedule"))
    if isinstance(loaded.get("employees"), list):
        result = {
            "employees": [_normalize_employee(e) for e in loaded["employees"]],
            "last_updated": loaded.get("last_updated"),
            "schedule": schedule,
        }
    else:
        employees = []
        for dept in (loaded.get("departments") or []):
            defects = _blank_defects()
            try:
                defects["off_shade"] = max(0, int(dept.get("errors", 0)))
            except (TypeError, ValueError):
                pass
            employees.append({"name": dept.get("name") or "Operator", "defects": defects})
        if not employees:
            employees = _sample_employees()
        result = {"employees": employees, "last_updated": loaded.get("last_updated"), "schedule": schedule}

    result.update(extra)
    return result


def load_state():
    if not os.path.exists(DATA_FILE):
        fresh = {"employees": _sample_employees(), "last_updated": None, "schedule": _default_schedule()}
        write_state(fresh)
        return fresh
    try:
        with open(DATA_FILE, "r", encoding="utf-8") as f:
            loaded = json.load(f)
        migrated = migrate_state(loaded)
        if migrated != loaded:
            write_state(migrated)
        return migrated
    except (json.JSONDecodeError, ValueError, OSError):
        fresh = {"employees": _sample_employees(), "last_updated": None, "schedule": _default_schedule()}
        write_state(fresh)
        return fresh


def write_state(snapshot):
    # Write to a temp file then atomically replace — prevents corrupt reads
    # if a reader opens the file mid-write.
    tmp = DATA_FILE + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(snapshot, f, indent=2)
    os.replace(tmp, DATA_FILE)


def employee_total(emp):
    return sum(emp["defects"].get(c["key"], 0) for c in DEFECT_CATEGORIES)


def get_state():
    """Always read from disk so changes made by the QC Control Center
    (or any other process writing data.json) are immediately visible."""
    with _data_lock:
        return load_state()


def persist(snapshot):
    """Write snapshot to disk. Call this while already holding _data_lock,
    or outside any lock for read-only callers like the scheduler.

    Before writing, re-reads whatever is currently on disk and folds in
    any top-level keys we don't recognize and didn't already have in
    `snapshot`. This guards against clobbering fields the QC Control
    Center admin tool wrote in the window between our get_state() read
    and this write — those extra keys are passed through untouched
    rather than disappearing on the next tray-side save.
    """
    try:
        with open(DATA_FILE, "r", encoding="utf-8") as f:
            on_disk = json.load(f)
        if isinstance(on_disk, dict):
            for k, v in on_disk.items():
                if k not in _KNOWN_TOP_KEYS and k not in snapshot:
                    snapshot[k] = v
    except (json.JSONDecodeError, ValueError, OSError):
        pass
    snapshot["last_updated"] = datetime.now().isoformat()
    write_state(snapshot)


# ==============================================================================
# FLASK APP
# ==============================================================================

app = Flask(__name__, static_folder=BASE_DIR)


@app.route("/")
def index():
    return send_from_directory(BASE_DIR, "dashboard.html")


@app.route("/logo.png")
def logo():
    try:
        return send_from_directory(BASE_DIR, "logo.png")
    except Exception:
        return ("", 204)


@app.route("/api/state")
def api_state():
    now  = datetime.now()
    hour = now.hour
    shift = "MORNING SHIFT" if 6 <= hour < 18 else "NIGHT SHIFT"

    s           = get_state()
    emps        = s["employees"]
    grand_total = sum(employee_total(e) for e in emps)

    raw = s.get("last_updated")
    if raw:
        try:
            dt       = datetime.fromisoformat(raw)
            last_upd = "© K.Calantog 2026, Last updated: " + dt.strftime("%I:%M:%S %p").lstrip("0")
        except ValueError:
            last_upd = "Last updated: --"
    else:
        last_upd = "© K.Calantog 2026, Last updated: --"

    return jsonify({
        "employees": [
            {"name": e["name"], "defects": e["defects"], "total": employee_total(e)}
            for e in emps
        ],
        "grand_total":      grand_total,
        "shift":            shift,
        "now":              {"date": now.strftime("%A, %B %#d, %Y"),
                             "time": now.strftime("%I:%M:%S %p").lstrip("0")},
        "last_updated":     last_upd,
        "categories":       DEFECT_CATEGORIES,
        "threshold_orange": THRESHOLD_ORANGE,
        "threshold_red":    THRESHOLD_RED,
        "schedule":         s.get("schedule", _default_schedule()),
    })


@app.route("/api/schedule", methods=["POST"])
def api_set_schedule():
    body = request.get_json(force=True, silent=True) or {}
    with _data_lock:
        s = get_state()
        s["schedule"] = _normalize_schedule(body)
        sched = s["schedule"]
        persist(s)
    return jsonify({"ok": True, "hour": sched["hour"], "minute": sched["minute"]})


@app.route("/api/login", methods=["POST"])
def api_login():
    body = request.get_json(force=True, silent=True) or {}
    if body.get("password") == ADMIN_PASSWORD:
        return jsonify({"ok": True})
    return jsonify({"ok": False, "error": "Incorrect password."}), 401


@app.route("/api/employees", methods=["POST"])
def api_employees_save():
    body     = request.get_json(force=True, silent=True) or {}
    raw_list = body.get("employees")
    if not isinstance(raw_list, list):
        return jsonify({"ok": False, "error": "Bad payload"}), 400
    with _data_lock:
        s = get_state()
        s["employees"] = [_normalize_employee(e) for e in raw_list]
        persist(s)
    return jsonify({"ok": True})


@app.route("/api/employees/add", methods=["POST"])
def api_add_employee():
    body = request.get_json(force=True, silent=True) or {}
    name = (body.get("name") or "").strip()
    if not name:
        return jsonify({"ok": False, "error": "Name required."}), 400
    with _data_lock:
        s = get_state()
        if any(e["name"].lower() == name.lower() for e in s["employees"]):
            return jsonify({"ok": False, "error": "Operator already exists."}), 409
        s["employees"].append({"name": name, "defects": _blank_defects()})
        persist(s)
    return jsonify({"ok": True})


@app.route("/api/employees/delete", methods=["POST"])
def api_delete_employee():
    body = request.get_json(force=True, silent=True) or {}
    name = body.get("name")
    with _data_lock:
        s = get_state()
        before = len(s["employees"])
        s["employees"] = [e for e in s["employees"] if e["name"] != name]
        if len(s["employees"]) == before:
            return jsonify({"ok": False, "error": "Not found."}), 404
        persist(s)
    return jsonify({"ok": True})


@app.route("/api/employees/rename", methods=["POST"])
def api_rename_employee():
    body     = request.get_json(force=True, silent=True) or {}
    old_name = body.get("old_name", "")
    new_name = (body.get("new_name") or "").strip()
    if not new_name:
        return jsonify({"ok": False, "error": "Name required."}), 400
    with _data_lock:
        s = get_state()
        for e in s["employees"]:
            if e["name"] == old_name:
                e["name"] = new_name
                persist(s)
                return jsonify({"ok": True})
    return jsonify({"ok": False, "error": "Not found."}), 404


@app.route("/api/employees/defect", methods=["POST"])
def api_set_defect():
    body  = request.get_json(force=True, silent=True) or {}
    name  = body.get("name")
    key   = body.get("key")
    value = body.get("value")
    try:
        value = max(0, int(value))
    except (TypeError, ValueError):
        return jsonify({"ok": False, "error": "Bad value"}), 400
    with _data_lock:
        s = get_state()
        for e in s["employees"]:
            if e["name"] == name:
                canonical_key = _KEY_LOOKUP.get(str(key).lower())
                if canonical_key is None:
                    return jsonify({"ok": False, "error": "Bad key"}), 400
                e["defects"][canonical_key] = value
                persist(s)
                return jsonify({"ok": True, "value": value})
    return jsonify({"ok": False, "error": "Not found."}), 404


@app.route("/api/employees/reset", methods=["POST"])
def api_reset_counters():
    with _data_lock:
        s = get_state()
        for e in s["employees"]:
            e["defects"] = _blank_defects()
        persist(s)
    return jsonify({"ok": True})


@app.route("/api/export/excel")
def api_export_excel():
    if not HAS_OPENPYXL:
        return jsonify({"ok": False, "error": "openpyxl not installed. Run: pip install openpyxl"}), 500

    wb = Workbook()
    ws = wb.active
    ws.title = "Defect Report"

    header = ["Operator"] + [c["label"] for c in DEFECT_CATEGORIES] + ["Total"]
    ws.append(header)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="101820", end_color="101820", fill_type="solid")
    for cell in ws[1]:
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center")

    emps = get_state()["employees"]
    for e in emps:
        ws.append([e["name"]] + [e["defects"].get(c["key"], 0) for c in DEFECT_CATEGORIES] + [employee_total(e)])

    grand_total = sum(employee_total(e) for e in emps)
    ws.append(["GRAND TOTAL"] + [""] * len(DEFECT_CATEGORIES) + [grand_total])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    for col_cells in ws.columns:
        values = [str(c.value) for c in col_cells if c.value is not None]
        width  = max([len(v) for v in values] + [10])
        ws.column_dimensions[col_cells[0].column_letter].width = width + 4

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = "defect_report_" + datetime.now().strftime("%Y%m%d_%H%M%S") + ".xlsx"
    resp = Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


def run_daily_export():
    """Write today's defect data into REPORTS_DIR/EMBRO-DEFECT-YY-MM.xlsx and reset
    counters. Shared by the manual 'Save Daily Snapshot' button, the manual
    /api/export/daily route, and the automatic server-side scheduler — so
    the behavior is identical no matter what triggered it.

    IMPORTANT: the workbook build + wb.save() happens OUTSIDE _data_lock.
    Saving to REPORTS_DIR can be slow or stall (network/shared-drive hiccup,
    antivirus scanning the file, someone has the .xlsx open in Excel) — if
    that I/O happened while holding _data_lock, every dashboard poll (which
    needs that same lock just to read in-memory state) would queue up
    behind it and the dashboard would freeze on "Connecting..." until the
    save finally finished. We only hold the lock for the brief, fast,
    in-memory snapshot and counter-reset steps, so a slow/stuck save can
    never block the dashboard.
    """
    if not HAS_OPENPYXL:
        return {"ok": False, "error": "openpyxl not installed. Run: pip install openpyxl"}

    reports_dir = REPORTS_DIR
    os.makedirs(reports_dir, exist_ok=True)

    now      = datetime.now()
    filename = "EMBRO-DEFECT-" + now.strftime("%y-%m") + ".xlsx"
    filepath = os.path.join(reports_dir, filename)
    sheet    = now.strftime("%d-%b").upper()   # e.g. "23-JUN"

    # ---- Snapshot current data — lock held only for this quick in-memory copy ----
    with _data_lock:
        s             = get_state()
        emps_snapshot = [{"name": e["name"], "defects": dict(e["defects"])} for e in s["employees"]]
    grand_total = sum(employee_total(e) for e in emps_snapshot)

    # ---- Build & save the workbook (slow/unreliable disk I/O — NO lock held) ----
    try:
        from openpyxl import load_workbook
        wb = load_workbook(filepath) if os.path.exists(filepath) else Workbook()
    except Exception:
        wb = Workbook()

    # Remove default blank sheet on fresh workbook
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
        del wb["Sheet"]

    # Avoid duplicate sheet names
    base  = sheet
    count = 1
    while sheet in wb.sheetnames:
        sheet = base + "_" + str(count)
        count += 1

    ws = wb.create_sheet(title=sheet)

    # Header row
    header      = ["Operator"] + [c["label"] for c in DEFECT_CATEGORIES] + ["Total"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="101820", end_color="101820", fill_type="solid")
    ws.append(header)
    for cell in ws[1]:
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for e in emps_snapshot:
        ws.append(
            [e["name"]] +
            [e["defects"].get(c["key"], 0) for c in DEFECT_CATEGORIES] +
            [employee_total(e)]
        )

    # Grand total row
    ws.append(["GRAND TOTAL"] + [""] * len(DEFECT_CATEGORIES) + [grand_total])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    # Column widths
    for col_cells in ws.columns:
        values = [str(c.value) for c in col_cells if c.value is not None]
        width  = max([len(v) for v in values] + [10])
        ws.column_dimensions[col_cells[0].column_letter].width = width + 4

    try:
        wb.save(filepath)
    except Exception as exc:
        # Most common causes: the .xlsx is currently open in Excel, or
        # reports_dir is a shared/network path that's unreachable right now.
        # Counters are deliberately left untouched so today's data isn't
        # lost — the next manual or scheduled export will retry.
        return {
            "ok": False,
            "error": f"Could not save report (is the file open in Excel? is {reports_dir} reachable?): {exc}",
        }

    # ---- Reset counters only after the save actually succeeded ----
    with _data_lock:
        s = get_state()
        for e in s["employees"]:
            e["defects"] = _blank_defects()
        persist(s)

    return {"ok": True, "file": filename, "sheet": sheet}


@app.route("/api/export/daily", methods=["POST"])
def api_export_daily():
    result = run_daily_export()
    return jsonify(result), (200 if result.get("ok") else 500)


# ==============================================================================
# FLASK SERVER THREAD  (stoppable via werkzeug)
# ==============================================================================

class FlaskServerThread(threading.Thread):
    """Runs Flask in a background thread; call .shutdown() to stop it."""

    def __init__(self):
        super().__init__(daemon=True, name="FlaskServer")
        self._server = None
        self._ready  = threading.Event()   # set once _server is assigned

    def run(self):
        try:
            self._server = make_server("0.0.0.0", PORT, app, threaded=True)
            self._ready.set()              # signal that _server is ready
            self._server.serve_forever()
        except Exception as exc:
            print(f"[FlaskServer] error: {exc}")
            self._ready.set()              # unblock shutdown even on error

    def shutdown(self):
        self._ready.wait(timeout=5)        # wait until _server is assigned
        if self._server:
            self._server.shutdown()
            self._server = None


# Global reference so tray callbacks can restart the server
_flask_thread: FlaskServerThread | None = None
_flask_lock   = threading.Lock()


def start_flask():
    global _flask_thread, _service_start_time
    with _flask_lock:
        if _flask_thread and _flask_thread.is_alive():
            return  # already running
        _flask_thread = FlaskServerThread()
        _flask_thread.start()
        _service_start_time = datetime.now()


def stop_flask():
    global _flask_thread, _service_start_time
    with _flask_lock:
        if _flask_thread:
            _flask_thread.shutdown()
            _flask_thread.join(timeout=5)
            _flask_thread = None
            _service_start_time = None


def flask_running() -> bool:
    return bool(_flask_thread and _flask_thread.is_alive())


# ==============================================================================
# AUTO-REPORT SCHEDULER  (server-side; runs as long as this app is running,
# regardless of whether any browser/PC has the dashboard open)
# ==============================================================================

_sched_fired_key  = ""    # "HH:MM" already fired this minute, to avoid double-firing
_sched_fired_lock = threading.Lock()
_sched_stop_event = threading.Event()
_sched_thread: threading.Thread | None = None


def _scheduler_loop():
    global _sched_fired_key
    while not _sched_stop_event.is_set():
        try:
            s     = get_state()
            sched = dict(s.get("schedule") or _default_schedule())
            hour, minute = sched.get("hour", -1), sched.get("minute", 0)
            if hour is not None and hour >= 0:
                now      = datetime.now()
                now_key  = f"{now.hour:02d}:{now.minute:02d}"
                hitting  = (now.hour == hour and now.minute == minute)
                with _sched_fired_lock:
                    already_fired = (_sched_fired_key == now_key)
                if hitting and not already_fired:
                    with _sched_fired_lock:
                        _sched_fired_key = now_key
                    try:
                        run_daily_export()
                    except Exception as exc:
                        print(f"[Scheduler] auto-export failed: {exc}")
                elif not hitting:
                    with _sched_fired_lock:
                        _sched_fired_key = ""
        except Exception as exc:
            print(f"[Scheduler] loop error: {exc}")
        # Poll every 10s — tight enough to never miss a minute, light on CPU
        _sched_stop_event.wait(10)


def start_scheduler():
    global _sched_thread
    if _sched_thread and _sched_thread.is_alive():
        return  # already running
    _sched_stop_event.clear()
    _sched_thread = threading.Thread(target=_scheduler_loop, daemon=True, name="ReportScheduler")
    _sched_thread.start()


def stop_scheduler():
    _sched_stop_event.set()


# ==============================================================================
# TRAY ICON  (16×16 or 32×32 PNG generated with Pillow)
# ==============================================================================

def _make_icon_image(size=64) -> Image.Image:
    """
    Draws a simple circular icon:
      • dark navy background
      • white 'D' letter  (Defect board)
      • small green dot in corner when service is running
    """
    img  = Image.new("RGBA", (size, size), (0, 0, 0, 0))
    draw = ImageDraw.Draw(img)

    # Background circle
    draw.ellipse([2, 2, size - 3, size - 3], fill="#101820")

    # Letter 'D'
    font_size = int(size * 0.52)
    try:
        font = ImageFont.truetype("arial.ttf", font_size)
    except IOError:
        font = ImageFont.load_default()

    text    = "E"
    bbox    = draw.textbbox((0, 0), text, font=font)
    tw, th  = bbox[2] - bbox[0], bbox[3] - bbox[1]
    tx      = (size - tw) // 2 - bbox[0]
    ty      = (size - th) // 2 - bbox[1]
    draw.text((tx, ty), text, fill="#FFD600", font=font)

    # Green status dot (bottom-right)
    dot_r = max(6, size // 8)
    draw.ellipse(
        [size - dot_r * 2 - 2, size - dot_r * 2 - 2, size - 2, size - 2],
        fill="#00E676", outline="#101820", width=1,
    )
    return img


def _make_stopped_icon_image(size=64) -> Image.Image:
    """Same as above but with a red dot to indicate service stopped."""
    img  = Image.new("RGBA", (size, size), (0, 0, 0, 0))
    draw = ImageDraw.Draw(img)
    draw.ellipse([2, 2, size - 3, size - 3], fill="#2a2a2a")

    font_size = int(size * 0.52)
    try:
        font = ImageFont.truetype("arial.ttf", font_size)
    except IOError:
        font = ImageFont.load_default()

    text    = "E"
    bbox    = draw.textbbox((0, 0), text, font=font)
    tw, th  = bbox[2] - bbox[0], bbox[3] - bbox[1]
    tx      = (size - tw) // 2 - bbox[0]
    ty      = (size - th) // 2 - bbox[1]
    draw.text((tx, ty), text, fill="#888888", font=font)

    dot_r = max(6, size // 8)
    draw.ellipse(
        [size - dot_r * 2 - 2, size - dot_r * 2 - 2, size - 2, size - 2],
        fill="#FF1744", outline="#2a2a2a", width=1,
    )
    return img


# ==============================================================================
# TRAY CALLBACKS
# ==============================================================================

def on_open_dashboard(icon, item):
    webbrowser.open(DASHBOARD_URL)


def on_open_reports(icon, item):
    """Open the reports folder in Windows Explorer."""
    os.makedirs(REPORTS_DIR, exist_ok=True)
    os.startfile(REPORTS_DIR)


def on_restart(icon, item):
    def _do_restart():
        icon.notify("Restarting service…", "Embroidery Defect Monitor")
        stop_flask()
        start_flask()
        icon.icon = _make_icon_image()
        icon.notify("Service restarted ✓", "Embroidery Defect Monitor")
    threading.Thread(target=_do_restart, daemon=True).start()


def on_stop_service(icon, item):
    def _do_stop():
        stop_flask()
        icon.icon = _make_stopped_icon_image()
        icon.notify("Service stopped. Right-click → Restart to bring it back.", "Embroidery Defect Monitor")
    threading.Thread(target=_do_stop, daemon=True).start()


def on_start_service(icon, item):
    def _do_start():
        if flask_running():
            icon.notify("Service is already running.", "Defect Monitor")
            return
        start_flask()
        icon.icon = _make_icon_image()
        icon.notify("Service started ✓", "Embroidery Defect Monitor")
    threading.Thread(target=_do_start, daemon=True).start()


def on_quit(icon, item):
    # Run teardown in a separate thread so the tray's event loop
    # is free to process icon.stop() without deadlocking.
    def _do_quit():
        stop_scheduler()
        stop_flask()
        icon.stop()
    threading.Thread(target=_do_quit, daemon=True).start()


def _format_uptime() -> str:
    """Return a human-readable uptime string, e.g. '2h 05m 30s'."""
    if _service_start_time is None:
        return ""
    delta   = datetime.now() - _service_start_time
    total_s = int(delta.total_seconds())
    h, rem  = divmod(total_s, 3600)
    m, s    = divmod(rem, 60)
    if h:
        return f"{h}h {m:02d}m {s:02d}s"
    elif m:
        return f"{m}m {s:02d}s"
    else:
        return f"{s}s"


def _service_status_text(icon):
    if flask_running():
        uptime = _format_uptime()
        uptime_str = f"  ·  up {uptime}" if uptime else ""
        return f"● Running on port {PORT}{uptime_str}"
    return "○ Stopped"


# ---------------------------------------------------------------------------
# Start-on-boot helpers (Windows Registry)
# ---------------------------------------------------------------------------

def _exe_path() -> str:
    """Return the path used to launch this process (exe or pythonw script)."""
    if getattr(sys, "frozen", False):          # PyInstaller bundle
        return f'"{sys.executable}"'
    # Running as a plain .py – use pythonw so no console appears on login
    pythonw = sys.executable.replace("python.exe", "pythonw.exe")
    script  = os.path.abspath(__file__)
    return f'"{pythonw}" "{script}"'


def _is_startup_enabled() -> bool:
    try:
        with winreg.OpenKey(winreg.HKEY_CURRENT_USER, _STARTUP_REG_KEY) as key:
            winreg.QueryValueEx(key, _STARTUP_APP_NAME)
            return True
    except FileNotFoundError:
        return False
    except OSError:
        return False


def _set_startup(enable: bool) -> None:
    with winreg.OpenKey(
        winreg.HKEY_CURRENT_USER, _STARTUP_REG_KEY,
        0, winreg.KEY_SET_VALUE
    ) as key:
        if enable:
            winreg.SetValueEx(key, _STARTUP_APP_NAME, 0, winreg.REG_SZ, _exe_path())
        else:
            try:
                winreg.DeleteValue(key, _STARTUP_APP_NAME)
            except FileNotFoundError:
                pass


def on_toggle_startup(icon, menu_item):
    """Toggle the start-on-boot registry entry and notify the user."""
    currently_enabled = _is_startup_enabled()
    try:
        _set_startup(not currently_enabled)
        if not currently_enabled:
            icon.notify("App will now start automatically on login.", "Embroidery Defect Monitor")
        else:
            icon.notify("Auto-start on login has been disabled.", "Embroidery Defect Monitor")
    except OSError as exc:
        icon.notify(f"Could not update startup setting:\n{exc}", "Embroidery Defect Monitor")


# ==============================================================================
# BUILD & RUN TRAY
# ==============================================================================

def _startup_label(icon):
    return "✔ Start on Boot (Login)" if _is_startup_enabled() else "    Start on Boot (Login)"


def build_menu() -> Menu:
    settings_submenu = Menu(
        item(_startup_label, on_toggle_startup),
    )

    return Menu(
        item("Embroidery Defect Monitoring Board", None, enabled=False),   # title row
        Menu.SEPARATOR,
        item("Open Dashboard",          on_open_dashboard),
        item("Open Reports (Excel)",    on_open_reports),
        Menu.SEPARATOR,
        item("Start Service",           on_start_service),
        item("Restart Service",         on_restart),
        item("Stop Service",            on_stop_service),
        Menu.SEPARATOR,
        item(_service_status_text,      None, enabled=False),   # live status + uptime
        Menu.SEPARATOR,
        item("Settings",                settings_submenu),
        Menu.SEPARATOR,
        item("Quit",                    on_quit),
    )


def main():
    # Start Flask before showing the tray
    start_flask()
    start_scheduler()

    tray_icon = pystray.Icon(
        name    = "EmbroideryDefectMonitor",
        icon    = _make_icon_image(),
        title   = "Embroidery Defect Monitoring Board",
        menu    = build_menu(),
    )

    tray_icon.run()   # blocks until icon.stop() is called


if __name__ == "__main__":
    main()
